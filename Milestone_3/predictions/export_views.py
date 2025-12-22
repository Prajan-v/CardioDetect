"""
Export views for CardioDetect.
Handles Excel and PDF exports.
"""

import io
from datetime import datetime
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework import status
from rest_framework.response import Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from .models import Prediction


class ExportPredictionsExcelView(APIView):
    """
    Export user's prediction history to Excel file.
    
    GET /api/history/export/excel/
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        # Get user's predictions
        predictions = Prediction.objects.filter(user=request.user).order_by('-created_at')
        
        if not predictions.exists():
            return Response({
                'error': 'No predictions to export'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Prediction History"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Headers
        headers = [
            "Date", "Time", "Input Method", "Risk Category", "Risk %",
            "Age", "Sex", "Systolic BP", "Diastolic BP", "Cholesterol",
            "HDL", "Glucose", "BMI", "Heart Rate", "Smoking", "Diabetes"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align
        
        # Data rows
        for row_idx, prediction in enumerate(predictions, 2):
            data = [
                prediction.created_at.strftime("%Y-%m-%d"),
                prediction.created_at.strftime("%H:%M:%S"),
                prediction.input_method.title() if prediction.input_method else "Manual",
                prediction.risk_category or "N/A",
                f"{prediction.risk_percentage:.1f}" if prediction.risk_percentage else "N/A",
                prediction.age or "",
                "M" if prediction.sex == 1 else "F" if prediction.sex == 0 else "",
                prediction.systolic_bp or "",
                prediction.diastolic_bp or "",
                prediction.cholesterol or "",
                prediction.hdl or "",
                prediction.glucose or "",
                f"{prediction.bmi:.1f}" if prediction.bmi else "",
                prediction.heart_rate or "",
                "Yes" if prediction.smoking else "No",
                "Yes" if prediction.diabetes else "No",
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border
                cell.alignment = center_align
                
                # Color code risk category
                if col == 4:  # Risk Category column
                    if value == "HIGH":
                        cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                        cell.font = Font(color="DC2626", bold=True)
                    elif value == "MODERATE":
                        cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
                        cell.font = Font(color="D97706", bold=True)
                    elif value == "LOW":
                        cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                        cell.font = Font(color="059669", bold=True)
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Add summary sheet
        ws_summary = wb.create_sheet("Summary")
        ws_summary.cell(row=1, column=1, value="CardioDetect Prediction Summary").font = Font(bold=True, size=14)
        ws_summary.cell(row=3, column=1, value=f"User: {request.user.get_full_name() or request.user.email}")
        ws_summary.cell(row=4, column=1, value=f"Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        ws_summary.cell(row=5, column=1, value=f"Total Predictions: {predictions.count()}")
        
        # Risk distribution
        high_count = predictions.filter(risk_category='HIGH').count()
        moderate_count = predictions.filter(risk_category='MODERATE').count()
        low_count = predictions.filter(risk_category='LOW').count()
        
        ws_summary.cell(row=7, column=1, value="Risk Distribution:").font = Font(bold=True)
        ws_summary.cell(row=8, column=1, value=f"  High Risk: {high_count}")
        ws_summary.cell(row=9, column=1, value=f"  Moderate Risk: {moderate_count}")
        ws_summary.cell(row=10, column=1, value=f"  Low Risk: {low_count}")
        
        # Save to response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"CardioDetect_Predictions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response


class ExportDoctorPatientsExcelView(APIView):
    """
    Export doctor's patient list to Excel file.
    
    GET /api/doctor/patients/export/excel/
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        # Check if user is a doctor
        if request.user.role != 'doctor':
            return Response({
                'error': 'Only doctors can export patient data'
            }, status=status.HTTP_403_FORBIDDEN)
        
        from .models import DoctorPatient
        
        # Get doctor's patients
        relationships = DoctorPatient.objects.filter(doctor=request.user).select_related('patient')
        
        if not relationships.exists():
            return Response({
                'error': 'No patients to export'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "My Patients"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Headers
        headers = [
            "Name", "Email", "Age", "Gender", "Phone",
            "Added Date", "Total Predictions", "Latest Risk", "Last Prediction"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align
        
        # Data rows
        for row_idx, rel in enumerate(relationships, 2):
            patient = rel.patient
            
            # Get patient's latest prediction
            latest_prediction = Prediction.objects.filter(user=patient).order_by('-created_at').first()
            total_predictions = Prediction.objects.filter(user=patient).count()
            
            data = [
                patient.get_full_name(),
                patient.email,
                patient.age or "",
                patient.get_gender_display() if patient.gender else "",
                patient.phone or "",
                rel.created_at.strftime("%Y-%m-%d") if rel.created_at else "",
                total_predictions,
                latest_prediction.risk_category if latest_prediction else "No predictions",
                latest_prediction.created_at.strftime("%Y-%m-%d") if latest_prediction else "",
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border
                cell.alignment = center_align
                
                # Color code risk category
                if col == 8 and latest_prediction:  # Latest Risk column
                    risk = latest_prediction.risk_category
                    if risk == "HIGH":
                        cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                        cell.font = Font(color="DC2626", bold=True)
                    elif risk == "MODERATE":
                        cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
                        cell.font = Font(color="D97706", bold=True)
                    elif risk == "LOW":
                        cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                        cell.font = Font(color="059669", bold=True)
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
        
        # Save to response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"CardioDetect_Patients_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
